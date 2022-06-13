# -*- coding: utf-8 -*-
"""
Created on Mon May 11 11:25:28 2022

@author: GC
* revised on 11.05.2022
"""
import pyblp
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy as sp
from scipy import stats
from openpyxl import load_workbook
import time

pyblp.options.digits = 4 #this is set at 2 in the tutorial
pyblp.options.verbose = False
pyblp.__version__


###############################################################################
# PT - Total costs
###############################################################################

#Total - off
###############################################################################

vfi = pd.read_csv(r'C:\Users\Gloria\Documents\1_Paper\Publication\Replication\R_paneldata\Output\i1rv.cvs', index_col=False)
product_data = vfi.loc[(vfi.block == "off"), ]
product_data = product_data.drop(columns=["nesting_ids"])

del product_data['Unnamed: 0']

product_data['tc'] = product_data.loc[:,['fcost', 'fcoa', 'rampc']].sum(axis=1)

X1 = pyblp.Formulation('0 + prices') 
X2 = pyblp.Formulation('0  + loadf') 
X3 = pyblp.Formulation('0 + tc') 

logit_formulations = (
    X1,
    X2,
    X3
   )
#logit_formulations

mc_integration = pyblp.Integration('monte_carlo', size=50, specification_options={'seed' : 0})
#mc_integration

problem = pyblp.Problem(logit_formulations, product_data, integration = mc_integration)
#problem

bfgs = pyblp.Optimization('bfgs', {'gtol' : 1e-10})
#bfgs

results = problem.solve(sigma=np.ones((1,1)),
                        beta = (0.5),
                        optimization = bfgs,
                        costs_bounds = (14.999, None))
results


elasticities = results.compute_elasticities
elasticities

#create directory inside python
import os
dirn = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\off"
os.makedirs(dirn)
os.chdir(dirn)

#Post-estimation results
import pickle
a = pyblp.ProblemResults.to_dict(results)
a.keys()
with open("offr.txt", "wb") as myFile:
    pickle.dump(a, myFile)

import pickle
with open("offr.txt", "rb") as myFile:
    offr = pickle.load(myFile)

#pvalues
a = offr
se = list([a[key] for key in ['gamma_se', 'beta_se', 'sigma_se']])
se = np.concatenate(se, axis=0 )

coeff = list([a[key] for key in ['gamma', 'beta', 'sigma']])
coeff = np.concatenate(coeff, axis=0 )
v = list([a[key] for key in ['xi']])
v = np.concatenate(v, axis=0 )
n=len(v)-5 # degrees of freedom n-p
ts= coeff/se

pval = (1-stats.t.cdf(np.abs(ts),n))*2 
interval = stats.norm.interval(0.68, loc=coeff, scale=se/np.sqrt(n))
CImin = interval[0]
CImax = interval[1]


#panda stats
index = ['tc', 'price', 'loadf']

df1 = pd.DataFrame(coeff, index=index, columns=['coeff'])
df2 = pd.DataFrame(se, index=index, columns=['se'])
df3 = pd.DataFrame(pval, index=index, columns=['p-val'])
df4 = pd.DataFrame(CImin, index= index, columns=['CImin'])
df5 = pd.DataFrame(CImax, index= index, columns=['CImax'])

offr= pd.concat([df1, df2, df3, df4, df5], axis=1)

#additional info
offr['Block']='off'
markets = list([a[key] for key in ['contraction_evaluations']])
markets = np.concatenate(markets, axis=0 )
T = len(markets)
offr['N']=T
obj = list([a[key] for key in ['objective']])
obj = int(obj[0])
offr['Obj']=obj
offr['Case']='tot'

with pd.ExcelWriter('stats28.xlsx') as writer:
    offr.to_excel(writer, sheet_name = 'pt', index = True)


#build an excel file that contains T market, prices, costs, shares, CS and PS
n2 = len(product_data)
stamp = pd.DataFrame(product_data, columns=['market_ids', 'block', 'hour', 'year', 'plant', 'prices', 'shares', 'tc'])
stamp.index = range(n2)
costs = results.compute_costs()
markups = results.compute_markups(costs=costs) #or here it is faster to use the previous calculation
profits = results.compute_profits(costs=costs)
cs = results.compute_consumer_surpluses() #esto estaba vacio
c= pd.DataFrame(costs, columns=['costs'])
cc = list([a[key] for key in ['clipped_costs']])
cc = pd.DataFrame(np.concatenate(cc, axis=0 ),columns=['clippedc'] )
d= pd.DataFrame(markups, columns=['mks'])
profitss=pd.DataFrame(profits, columns=['ps'])
consumers=pd.DataFrame(cs, columns=['cs']) 

#merging
stampt = stamp[0:len(markets)]
merge = pd.concat([stamp, c,cc, d , profitss], axis=1)
merge['Case']='tot'
x = pd.DataFrame(stamp['market_ids'].unique())
merge2 = pd.concat([x, consumers], axis=1)
merge2['Block']='off'
merge2['Case']='tot'

path = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\off\stats28.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
merge.to_excel(writer, sheet_name = 'one', index = True)
merge2.to_excel(writer, sheet_name = 'two', index = True)
writer.save()


#curvature
elasticities = results.compute_elasticities() 

aggregates = results.compute_aggregate_elasticities(factor=0.1)
aggregates2 = results.compute_aggregate_elasticities(factor=0.2)
aggregates3 = results.compute_aggregate_elasticities(factor=0.3)
aggregates4 = results.compute_aggregate_elasticities(factor=0.4)
aggregates5 = results.compute_aggregate_elasticities(factor=0.5)
aggregates6 = results.compute_aggregate_elasticities(factor=0.6)
aggregates7 = results.compute_aggregate_elasticities(factor=0.7)
aggregates8 = results.compute_aggregate_elasticities(factor=0.8)

agg = pd.DataFrame(aggregates,columns=['agg0.1'] )
agg2 = pd.DataFrame(aggregates2,columns=['agg0.2'] )
agg3 = pd.DataFrame(aggregates3,columns=['agg0.3'] )
agg4 = pd.DataFrame(aggregates4,columns=['agg0.4'] )
agg5 = pd.DataFrame(aggregates5,columns=['agg0.5'] )
agg6 = pd.DataFrame(aggregates6,columns=['agg0.6'] )
agg7 = pd.DataFrame(aggregates7,columns=['agg0.7'] )
agg8 = pd.DataFrame(aggregates8,columns=['agg0.8'] )

curvature = pd.concat([agg,agg2,agg3,agg4,agg5,agg6,agg7,agg8 ], axis=1, sort=False)

path = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\off\stats28.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
curvature.to_excel(writer, sheet_name = 'curv', index = True)
writer.save()

#additional data
diversions = results.compute_diversion_ratios()
means = results.extract_diagonal_means(elasticities)
e=pd.DataFrame(elasticities)
di=pd.DataFrame(diversions) 
ie1_o = pd.concat([e, di], axis=1, sort=False)
ie1_o.to_csv('i_1.csv', index = False)



#Total - p1
###############################################################################

vfi = pd.read_csv(r'C:\Users\Gloria\Documents\1_Paper\Publication\Replication\R_paneldata\Output\i1rv.cvs', index_col=False)
product_data = vfi.loc[(vfi.block == "peak1"), ]
product_data = product_data.drop(columns=["nesting_ids"])

del product_data['Unnamed: 0']

product_data['tc'] = product_data.loc[:,['fcost', 'fcoa', 'rampc']].sum(axis=1)

X1 = pyblp.Formulation('0 + prices') 
X2 = pyblp.Formulation('0  + loadf') 
X3 = pyblp.Formulation('0 + tc') 

logit_formulations = (
    X1,
    X2,
    X3
   )
#logit_formulations

mc_integration = pyblp.Integration('monte_carlo', size=50, specification_options={'seed' : 0})
#mc_integration

problem = pyblp.Problem(logit_formulations, product_data, integration = mc_integration)
#problem

bfgs = pyblp.Optimization('bfgs', {'gtol' : 1e-10})
#bfgs

results = problem.solve(sigma=np.ones((1,1)),
                        beta = (0.5),
                        optimization = bfgs,
                        costs_bounds = (14.999, None))
results


elasticities = results.compute_elasticities
elasticities


#create directory inside python
import os
dirn = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\p1"
os.makedirs(dirn)
os.chdir(dirn)

import pickle
a = pyblp.ProblemResults.to_dict(results)
a.keys()
with open("p1r.txt", "wb") as myFile:
    pickle.dump(a, myFile)

#Post-estimation results
import pickle
with open("p1r.txt", "rb") as myFile:
    p1r = pickle.load(myFile)

#pvalues
a = p1r
se = list([a[key] for key in ['gamma_se', 'beta_se', 'sigma_se']])
se = np.concatenate(se, axis=0 )

coeff = list([a[key] for key in ['gamma', 'beta', 'sigma']])
coeff = np.concatenate(coeff, axis=0 )
v = list([a[key] for key in ['xi']])
v = np.concatenate(v, axis=0 )
n=len(v)-5 # degrees of freedom n-p
ts= coeff/se

pval = (1-stats.t.cdf(np.abs(ts),n))*2 
interval = stats.norm.interval(0.68, loc=coeff, scale=se/np.sqrt(n))
CImin = interval[0]
CImax = interval[1]


#panda stats
index = ['tc', 'price', 'loadf']

df1 = pd.DataFrame(coeff, index=index, columns=['coeff'])
df2 = pd.DataFrame(se, index=index, columns=['se'])
df3 = pd.DataFrame(pval, index=index, columns=['p-val'])
df4 = pd.DataFrame(CImin, index= index, columns=['CImin'])
df5 = pd.DataFrame(CImax, index= index, columns=['CImax'])

p1r= pd.concat([df1, df2, df3, df4, df5], axis=1)

#additional info
p1r['Block']='peak1'
markets = list([a[key] for key in ['contraction_evaluations']])
markets = np.concatenate(markets, axis=0 )
T = len(markets)
p1r['N']=T
obj = list([a[key] for key in ['objective']])
obj = int(obj[0])
p1r['Obj']=obj
p1r['Case']='tot'


with pd.ExcelWriter('stats29.xlsx') as writer:
    p1r.to_excel(writer, sheet_name = 'pt', index = True)


#build an excel file that contains T market, prices, costs, shares, CS and PS
n2 = len(product_data)
stamp = pd.DataFrame(product_data, columns=['market_ids', 'block', 'hour', 'year', 'plant', 'prices', 'shares', 'tc'])
stamp.index = range(n2)
costs = results.compute_costs()
markups = results.compute_markups(costs=costs) #or here it is faster to use the previous calculation
profits = results.compute_profits(costs=costs)
cs = results.compute_consumer_surpluses() #esto estaba vacio
c= pd.DataFrame(costs, columns=['costs'])
cc = list([a[key] for key in ['clipped_costs']])
cc = pd.DataFrame(np.concatenate(cc, axis=0 ),columns=['clippedc'] )
d= pd.DataFrame(markups, columns=['mks'])
profitss=pd.DataFrame(profits, columns=['ps'])
consumers=pd.DataFrame(cs, columns=['cs']) 

#merging
stampt = stamp[0:len(markets)]
merge = pd.concat([stamp, c,cc, d , profitss], axis=1)
merge['Case']='tot'
x = pd.DataFrame(stamp['market_ids'].unique())
merge2 = pd.concat([x, consumers], axis=1)
merge2['Block']='peak1'
merge2['Case']='tot'

path = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\p1\stats29.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
merge.to_excel(writer, sheet_name = 'one', index = True)
merge2.to_excel(writer, sheet_name = 'two', index = True)
writer.save()


#curvature
elasticities = results.compute_elasticities() 

aggregates = results.compute_aggregate_elasticities(factor=0.1)
aggregates2 = results.compute_aggregate_elasticities(factor=0.2)
aggregates3 = results.compute_aggregate_elasticities(factor=0.3)
aggregates4 = results.compute_aggregate_elasticities(factor=0.4)
aggregates5 = results.compute_aggregate_elasticities(factor=0.5)
aggregates6 = results.compute_aggregate_elasticities(factor=0.6)
aggregates7 = results.compute_aggregate_elasticities(factor=0.7)
aggregates8 = results.compute_aggregate_elasticities(factor=0.8)

agg = pd.DataFrame(aggregates,columns=['agg0.1'] )
agg2 = pd.DataFrame(aggregates2,columns=['agg0.2'] )
agg3 = pd.DataFrame(aggregates3,columns=['agg0.3'] )
agg4 = pd.DataFrame(aggregates4,columns=['agg0.4'] )
agg5 = pd.DataFrame(aggregates5,columns=['agg0.5'] )
agg6 = pd.DataFrame(aggregates6,columns=['agg0.6'] )
agg7 = pd.DataFrame(aggregates7,columns=['agg0.7'] )
agg8 = pd.DataFrame(aggregates8,columns=['agg0.8'] )

curvature = pd.concat([agg,agg2,agg3,agg4,agg5,agg6,agg7,agg8 ], axis=1, sort=False)

path = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\p1\stats29.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
curvature.to_excel(writer, sheet_name = 'curv', index = True)
writer.save()

#additional data
diversions = results.compute_diversion_ratios()
means = results.extract_diagonal_means(elasticities)
e=pd.DataFrame(elasticities)
di=pd.DataFrame(diversions) 
ie1_o = pd.concat([e, di], axis=1, sort=False)
ie1_o.to_csv('i_1.csv', index = False)


#Total - p2
###############################################################################

vfi = pd.read_csv(r'C:\Users\Gloria\Documents\1_Paper\Publication\Replication\R_paneldata\Output\i1rv.cvs', index_col=False)
product_data = vfi.loc[(vfi.block == "peak2"), ]
product_data = product_data.drop(columns=["nesting_ids"])

del product_data['Unnamed: 0']

product_data['tc'] = product_data.loc[:,['fcost', 'fcoa', 'rampc']].sum(axis=1)

X1 = pyblp.Formulation('0 + prices') 
X2 = pyblp.Formulation('0  + loadf') 
X3 = pyblp.Formulation('0 + tc') 

logit_formulations = (
    X1,
    X2,
    X3
   )
#logit_formulations

mc_integration = pyblp.Integration('monte_carlo', size=50, specification_options={'seed' : 0})
#mc_integration

problem = pyblp.Problem(logit_formulations, product_data, integration = mc_integration)
#problem

bfgs = pyblp.Optimization('bfgs', {'gtol' : 1e-10})
#bfgs

results = problem.solve(sigma=np.ones((1,1)),
                        beta = (0.5),
                        optimization = bfgs,
                        costs_bounds = (14.999, None))
results


elasticities = results.compute_elasticities
elasticities


#create directory inside python
import os
dirn = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\p2"
os.makedirs(dirn)
os.chdir(dirn)

#Post-estimation results
import pickle
a = pyblp.ProblemResults.to_dict(results)
a.keys()
with open("p2r.txt", "wb") as myFile:
    pickle.dump(a, myFile)

import pickle
with open("p2r.txt", "rb") as myFile:
    p2r = pickle.load(myFile)

#pvalues
a = p2r
se = list([a[key] for key in ['gamma_se', 'beta_se', 'sigma_se']])
se = np.concatenate(se, axis=0 )

coeff = list([a[key] for key in ['gamma', 'beta', 'sigma']])
coeff = np.concatenate(coeff, axis=0 )
v = list([a[key] for key in ['xi']])
v = np.concatenate(v, axis=0 )
n=len(v)-5 # degrees of freedom n-p
ts= coeff/se

pval = (1-stats.t.cdf(np.abs(ts),n))*2 
interval = stats.norm.interval(0.68, loc=coeff, scale=se/np.sqrt(n))
CImin = interval[0]
CImax = interval[1]


#panda stats
index = ['tc', 'price', 'loadf']

df1 = pd.DataFrame(coeff, index=index, columns=['coeff'])
df2 = pd.DataFrame(se, index=index, columns=['se'])
df3 = pd.DataFrame(pval, index=index, columns=['p-val'])
df4 = pd.DataFrame(CImin, index= index, columns=['CImin'])
df5 = pd.DataFrame(CImax, index= index, columns=['CImax'])

p2r= pd.concat([df1, df2, df3, df4, df5], axis=1)

#additional info
p2r['Block']='peak2'
markets = list([a[key] for key in ['contraction_evaluations']])
markets = np.concatenate(markets, axis=0 )
T = len(markets)
p2r['N']=T
obj = list([a[key] for key in ['objective']])
obj = int(obj[0])
p2r['Obj']=obj
p2r['Case']='tot'

with pd.ExcelWriter('stats30.xlsx') as writer:
    p2r.to_excel(writer, sheet_name = 'pt', index = True)



#build an excel file that contains T market, prices, costs, shares, CS and PS
n2 = len(product_data)
stamp = pd.DataFrame(product_data, columns=['market_ids', 'block', 'hour', 'year', 'plant', 'prices', 'shares', 'tc'])
stamp.index = range(n2)
costs = results.compute_costs()
markups = results.compute_markups(costs=costs) #or here it is faster to use the previous calculation
profits = results.compute_profits(costs=costs)
cs = results.compute_consumer_surpluses() #esto estaba vacio
c= pd.DataFrame(costs, columns=['costs'])
cc = list([a[key] for key in ['clipped_costs']])
cc = pd.DataFrame(np.concatenate(cc, axis=0 ),columns=['clippedc'] )
d= pd.DataFrame(markups, columns=['mks'])
profitss=pd.DataFrame(profits, columns=['ps'])
consumers=pd.DataFrame(cs, columns=['cs']) 

#merging
stampt = stamp[0:len(markets)]
merge = pd.concat([stamp, c,cc, d , profitss], axis=1)
merge['Case']='tot'
x = pd.DataFrame(stamp['market_ids'].unique())
merge2 = pd.concat([x, consumers], axis=1)
merge2['Block']='peak2'
merge2['Case']='tot'

path = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\p2\stats30.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
merge.to_excel(writer, sheet_name = 'one', index = True)
merge2.to_excel(writer, sheet_name = 'two', index = True)
writer.save()
writer.close()

#curvature
elasticities = results.compute_elasticities() 

aggregates = results.compute_aggregate_elasticities(factor=0.1)
aggregates2 = results.compute_aggregate_elasticities(factor=0.2)
aggregates3 = results.compute_aggregate_elasticities(factor=0.3)
aggregates4 = results.compute_aggregate_elasticities(factor=0.4)
aggregates5 = results.compute_aggregate_elasticities(factor=0.5)
aggregates6 = results.compute_aggregate_elasticities(factor=0.6)
aggregates7 = results.compute_aggregate_elasticities(factor=0.7)
aggregates8 = results.compute_aggregate_elasticities(factor=0.8)

agg = pd.DataFrame(aggregates,columns=['agg0.1'] )
agg2 = pd.DataFrame(aggregates2,columns=['agg0.2'] )
agg3 = pd.DataFrame(aggregates3,columns=['agg0.3'] )
agg4 = pd.DataFrame(aggregates4,columns=['agg0.4'] )
agg5 = pd.DataFrame(aggregates5,columns=['agg0.5'] )
agg6 = pd.DataFrame(aggregates6,columns=['agg0.6'] )
agg7 = pd.DataFrame(aggregates7,columns=['agg0.7'] )
agg8 = pd.DataFrame(aggregates8,columns=['agg0.8'] )

curvature = pd.concat([agg,agg2,agg3,agg4,agg5,agg6,agg7,agg8 ], axis=1, sort=False)

path = r"C:\Users\Gloria\Documents\1_Paper\Publication\Replication\Python\Total\p2\stats30.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
curvature.to_excel(writer, sheet_name = 'curv', index = True)
writer.save()

#additional data
diversions = results.compute_diversion_ratios()
means = results.extract_diagonal_means(elasticities)
e=pd.DataFrame(elasticities)
di=pd.DataFrame(diversions) 
ie1_o = pd.concat([e, di], axis=1, sort=False)
ie1_o.to_csv('i_1.csv', index = False)

